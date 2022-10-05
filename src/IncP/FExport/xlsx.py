import sys
import logging
#--- xlsx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.comments import Comment
#
from Inc.DB.DbList import TDbList, TDbRec
from Inc.Util.UObj import GetNestedKey
from .FExport import TFExport


class TFExport_xlsx(TFExport):
    def Save(self, aDBl: list, aMatch: dict):
        logging.info('Result: xlsx')

        MatchName = 'Match'
        HeadCol = {'Mpn': 1, 'Name': 2, 'Price': 3, 'Match': 4}
        HeadRow = {'Title': 1, 'Data': 2}

        WB = Workbook()
        Sheet = WB['Sheet']
        WB.remove(Sheet)

        #--- match sheet
        WS = WB.create_sheet(title=MatchName)

        ConfType = GetNestedKey(self.Conf, 'Type.xlsx')
        ConfHeadLen = ConfType['HeadLen']
        ConfFile = ConfType['File']
        ConfLinks = ConfType['Links']
        ConfFormat = ConfType['Format']
        ConfRatio = ConfType['Ratio']
        logging.info('Create %s'  % (ConfFile))

        ColDim = WS.column_dimensions
        ColDim[get_column_letter(HeadCol['Price'])].font = Font(bold=True)
        ColDim[get_column_letter(HeadCol['Price'])].number_format = ConfFormat
        for x in ['Mpn', 'Name']:
            ColDim[get_column_letter(HeadCol[x])].width = ConfHeadLen[x]

        for x in ['Mpn', 'Name', 'Price']:
            WS.cell(HeadRow['Title'], HeadCol[x]).value = x

        for ColIdx, DBl in enumerate(aDBl, HeadCol['Match']):
            ColDim[get_column_letter(ColIdx)].number_format = ConfFormat
            WS.cell(HeadRow['Title'], ColIdx).value = DBl.Tag

        WS.freeze_panes = WS.cell(HeadRow['Data'], HeadCol['Match'])

        logging.info('Sheet: %s, Records %s' % (MatchName, len(aMatch)))
        for MatchNo, (MatchKey, MatchVal) in enumerate(aMatch.items(), 2):
            WS.cell(MatchNo, HeadCol['Mpn']).value = MatchKey
            PriceMin = sys.maxsize
            FileNoMin = -1
            Prices = []
            for FileNo, RecNo in MatchVal:
                DBl: TDbList = aDBl[FileNo]
                Rec: TDbRec = DBl.RecGo(RecNo)

                Price = Rec.GetField('Price')
                Prices.append(Price)
                if (Price) and (Price < PriceMin):
                    PriceMin = Price
                    FileNoMin = FileNo

                Cell = WS.cell(MatchNo, HeadCol['Match'] + FileNo)
                Cell.value = Price
                Cell.number_format = ConfFormat

                if (ConfLinks):
                   Cell.hyperlink = '%s#%s!A%s' % (ConfFile, DBl.Tag, RecNo + HeadRow['Data'])
            WS.cell(MatchNo, HeadCol['Price']).value = PriceMin
            WS.cell(MatchNo, HeadCol['Name']).value = Rec.GetField('Name')
            WS.cell(MatchNo, HeadCol['Match'] + FileNoMin).font = Font(bold=True)

            if (ConfRatio):
                for PriceNo, Price in enumerate(Prices):
                    if (Price) and (Price != PriceMin):
                        Text = '+%0.2f %0.2f%%' % (Price - PriceMin, Price / PriceMin * 100 - 100)
                        WS.cell(MatchNo, HeadCol['Match'] + PriceNo).comment = Comment(Text, '', 20)

        #--- vendors sheet
        for DBl in aDBl:
            Vendor = DBl.Tag
            logging.info('Sheet: %s' % Vendor)
            WS = WB.create_sheet(title=Vendor)
            WS.freeze_panes = WS.cell(HeadRow['Data'], 1)
            ColDim = WS.column_dimensions

            ConfVendor = self.Parent.Conf.GetVendor(Vendor)
            ConfFields = ConfVendor['Fields']
            ConfFieldsList = list(ConfFields.keys())
            ConfSheet = ConfVendor.get('Sheet', 'TDSheet')
            for FieldNo, Field in enumerate(ConfFields):
                TitlePrice = Field
                if (Field == 'Price'):
                    TitlePrice = '%s (%s)' % (Field, ConfVendor['USD'])
                    ColDim[get_column_letter(FieldNo + 1)].number_format = ConfFormat

                WS.cell(HeadRow['Title'], FieldNo + 1).value = TitlePrice

                Width = ConfHeadLen.get(Field)
                if (Width):
                    ColDim[get_column_letter(FieldNo+1)].width = Width

            MatchList = list(aMatch.keys())
            for RecNo, Rec in enumerate(DBl):
                Row = RecNo + HeadRow['Data']

                Mpn = Rec.GetField('Mpn')
                if (aMatch.get(Mpn)):
                    Cell = WS.cell(Row, ConfFieldsList.index('Mpn') + 1)
                    Cell.font = Font(bold=True)
                    if (ConfLinks):
                        Cell.hyperlink = '%s#%s!A%s' % (ConfFile, MatchName, MatchList.index(Mpn) + HeadRow['Data'])
                        if ('.xls' in Vendor):
                            WS.cell(Row, ConfFieldsList.index('Name') + 1).hyperlink = '%s#%s!A%s' % (Vendor, ConfSheet, Rec.GetField('No') + 1)

                for FieldNo, Field in enumerate(ConfFields):
                    Cell = WS.cell(Row, FieldNo + 1)
                    Cell.value = Rec.GetField(Field)

        logging.info('Save %s'  % (ConfFile))
        WB.save(ConfFile)
